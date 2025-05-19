import os
import subprocess
from openpyxl import load_workbook

def main():
    # 获取当前工作目录
    base_dir = os.getcwd()

    # 读取 Excel 文件
    workbook = load_workbook('apps_config/apps.xlsx')
    sheet = workbook.active

    # 获取表头
    headers = [cell.value for cell in sheet[1]]

    # 遍历每一行（从第二行开始）
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))
        app_name = row_data.get('app_name')
        short_id = row_data.get('short_id')
        display_name = row_data.get('display_name')
        resource_folder = row_data.get('resource_folder')

        # 构建项目路径
        project_path = os.path.join(base_dir, "generated_apps", app_name)

        # 检查项目路径是否存在
        if not os.path.isdir(project_path):
            print(f"项目目录 {project_path} 不存在，跳过该应用。")
            continue

        # 切换到项目目录
        os.chdir(project_path)

        print(f"开始上传 {app_name}...")

        # 使用 subprocess.Popen 实时打印输出
        process = subprocess.Popen(
            ['fastlane', 'build_and_upload'],
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True
        )

        # 实时读取并打印输出
        for line in process.stdout:
            print(line, end='')

        process.wait()

        if process.returncode == 0:
            print(f"{app_name} 上传成功。")
        else:
            print(f"{app_name} 上传失败，返回码：{process.returncode}")

        # 返回上级目录
        os.chdir(base_dir)

if __name__ == '__main__':
    main()
