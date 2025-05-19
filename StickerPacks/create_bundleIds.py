from openpyxl import load_workbook
import subprocess

# 配置文件路径
CONFIG_FILE = 'apps_config/apps.xlsx'

def create_bundle_ids():
    # 加载 Excel 文件
    wb = load_workbook(CONFIG_FILE)
    ws = wb.active  # 默认读取活动工作表

    # 获取标题行
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    # 确保包含所需的列
    if 'short_id' not in headers or 'app_name' not in headers:
        print("❌ Excel 文件中缺少 'short_id' 或 'app_name' 列。")
        return

    short_id_idx = headers.index('short_id')
    app_name_idx = headers.index('app_name')

    for row in ws.iter_rows(min_row=2, values_only=True):
        short_id = str(row[short_id_idx]).strip()
        app_name = str(row[app_name_idx]).strip()
        bundle_id = f"com.getsticker.stickerpack.{short_id}"

        print(f"🔧 正在创建 Bundle ID: {bundle_id}")

        try:
            subprocess.run([
                "fastlane", "produce",
                "--username", "192938268@qq.com",
                "--app_identifier", bundle_id,
                "--app_name", app_name,
                "--sku", bundle_id,
                "--app_version", "1.0",
                "--language", "English",
                "--itc_team_id", "1203961"
            ], check=True)
            print(f"✅ 成功创建: {bundle_id}")
        except subprocess.CalledProcessError as e:
            error_output = e.stderr.decode() if e.stderr else str(e)
            if "already exists" in error_output:
                print(f"⚠️  Bundle ID 已存在，跳过：{bundle_id}")
            else:
                print(f"❌ 创建失败：{bundle_id}，错误信息：{error_output}")

if __name__ == '__main__':
    create_bundle_ids()
